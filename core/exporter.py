import io
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# EXPORT — EXCEL
# ──────────────────────────────────────────────────────────────────────────────

def export_excel(df: pd.DataFrame, jd: dict) -> bytes:
    """
    Build a colour-coded Excel workbook with two sheets:
      Sheet 1 — Ranked Candidates (green ≥70 · yellow ≥45 · red <45)
      Sheet 2 — JD Summary

    Returns:
        bytes: Raw .xlsx file content ready for st.download_button.
    """
    wb = Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_fill  = PatternFill("solid", fgColor="6366F1")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin",   color="D1D5DB"),
        right=Side(style="thin",  color="D1D5DB"),
        top=Side(style="thin",    color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    fill_green  = PatternFill("solid", fgColor="D1FAE5")  # score ≥ 70
    fill_yellow = PatternFill("solid", fgColor="FEF3C7")  # score ≥ 45
    fill_red    = PatternFill("solid", fgColor="FEE2E2")  # score < 45

    # ── Sheet 1: Ranked Candidates ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Ranked Candidates"

    # Title banner
    ws.merge_cells("A1:M1")
    ws["A1"] = f"🎯 AI Recruiter — Results for: {jd.get('role', 'Role')}"
    ws["A1"].font  = Font(bold=True, size=14, color="1E293B", name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill  = PatternFill("solid", fgColor="F1F5F9")
    ws.row_dimensions[1].height = 30

    headers = [
        "Rank", "Name", "Role", "Location", "Experience (yrs)",
        "Skills", "Total Score", "Skill Score", "Role Score",
        "Exp Score", "Location Match", "Matched Skills", "Missing Skills",
    ]
    col_widths = [6, 22, 22, 14, 16, 32, 13, 13, 12, 12, 16, 30, 30]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[2].height = 24

    # Map logical keys → actual df column names (graceful fallback)
    key_map = [
        "rank", "name", "role", "location", "experience",
        "skills", "total_score", "skill_score", "role_score",
        "experience_score", "location_match", "matched_skills", "missing_skills",
    ]
    df_cols = set(df.columns)

    def _get(row, key):
        return row[key] if key in df_cols else ""

    for r_idx, (_, row) in enumerate(df.iterrows(), start=3):
        total = _get(row, "total_score")
        row_fill = fill_green if total >= 70 else (fill_yellow if total >= 45 else fill_red)

        for c_idx, key in enumerate(key_map, start=1):
            val  = _get(row, key)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if key == "total_score":
                cell.fill = row_fill
                cell.font = Font(bold=True, name="Calibri", size=11)
            elif key == "rank" and val in (1, 2, 3):
                cell.font = Font(bold=True, color="4F46E5", name="Calibri")

        ws.row_dimensions[r_idx].height = 18

    ws.freeze_panes = "A3"

    # ── Sheet 2: JD Summary ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("JD Summary")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 55

    ws2.merge_cells("A1:B1")
    ws2["A1"] = "📋 Job Description — Extracted Details"
    ws2["A1"].font      = Font(bold=True, size=13, color="1E293B", name="Calibri")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["A1"].fill      = PatternFill("solid", fgColor="EDE9FE")
    ws2.row_dimensions[1].height = 28

    jd_fields = [
        ("Role",            jd.get("role", "")),
        ("Location",        jd.get("location", "")),
        ("Experience Min",  f"{jd.get('experience_min', '')} years"),
        ("Experience Max",  f"{jd.get('experience_max', '')} years"),
        ("Industry",        jd.get("industry", "")),
        ("Employment Type", jd.get("employment_type", "")),
        ("Education",       jd.get("education", "") or "Not specified"),
        ("Company",         jd.get("company", "")  or "Not specified"),
        ("Required Skills", ", ".join(jd.get("skills", []))),
        ("Summary",         jd.get("summary", "")),
    ]

    lbl_font = Font(bold=True, name="Calibri", color="4F46E5")
    for i, (label, value) in enumerate(jd_fields, start=2):
        ws2.cell(row=i, column=1, value=label).font = lbl_font
        ws2.cell(row=i, column=2, value=value)
        ws2.row_dimensions[i].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# EXPORT — CSV
# ──────────────────────────────────────────────────────────────────────────────

def export_csv(df: pd.DataFrame) -> bytes:
    """Return the scored DataFrame as UTF-8 CSV bytes."""
    return df.to_csv(index=False).encode("utf-8")


# ──────────────────────────────────────────────────────────────────────────────
# QUICK SELF-TEST  (run: python filter.py)
# ──────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import io as _io

    sample_csv = """name,role,location,experience,skills
Priya Sharma,Data Scientist,Mumbai,4,Python ML SQL Pandas Sklearn
Rohit Verma,ML Engineer,Mumbai,3,Python ML TensorFlow NLP
Arjun Rao,Business Analyst,Delhi,5,Excel SQL Tableau
"""
    sample_jd = {
        "role": "Data Scientist",
        "skills": ["python", "ml", "sql", "pandas"],
        "location": "mumbai",
        "experience_min": 3,
        "experience_max": 6,
    }

    df_in = pd.read_csv(_io.StringIO(sample_csv))
    df_in.columns = [c.strip().lower().replace(" ", "_") for c in df_in.columns]
    scored, name_col = score_candidates(df_in, sample_jd)
    print(scored[["rank", "name", "total_score", "skill_score", "role_score", "experience_score"]].to_string(index=False))
    print(f"\nName column detected: {name_col}")
