"""
filter.py
─────────
Handles everything after the JD is parsed:
  1. load_candidates()     — reads a CSV and normalises column names
  2. score_candidates()    — filters & scores each candidate against the JD
  3. export_excel()        — produces a formatted, colour-coded .xlsx report
  4. export_csv()          — returns scored results as CSV bytes

Scoring formula
───────────────
  total_score (0–100) = skill_score (0–40)
                      + role_score  (0–30)
                      + exp_score   (0–30)

Skill score  : (# required skills found in candidate profile) / (# required) × 40
Role score   : keyword overlap between JD title and candidate title / JD words × 30
Exp score    : 30 if within [exp_min, exp_max]; reduced linearly for over/under-qualified
"""

import io
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────────────────────────────────────
# COLUMN NAME VARIANTS  (add aliases here if your CSV uses different headings)
# ──────────────────────────────────────────────────────────────────────────────
_COL_ALIASES: dict[str, list[str]] = {
    "name":       ["name", "candidate_name", "full_name", "applicant"],
    "role":       ["role", "job_title", "title", "position", "designation"],
    "location":   ["location", "city", "place", "loc", "region"],
    "experience": ["experience", "exp", "years_of_experience", "years", "yoe"],
    "skills":     ["skills", "skill", "tech_skills", "technical_skills", "technologies"],
    "company":    ["company", "current_company", "employer", "organisation", "organization"],
    "education":  ["education", "edu", "degree", "qualification"],
}


def _find_col(df: pd.DataFrame, field: str) -> str | None:
    """Return the first DataFrame column that matches any alias for `field`."""
    for alias in _COL_ALIASES.get(field, []):
        if alias in df.columns:
            return alias
    return None


def _parse_experience(raw) -> float:
    """Extract the first numeric value from an experience cell."""
    nums = re.findall(r"\d+\.?\d*", str(raw))
    return float(nums[0]) if nums else 0.0


# ──────────────────────────────────────────────────────────────────────────────
# LOAD
# ──────────────────────────────────────────────────────────────────────────────

def load_candidates(uploaded_file) -> pd.DataFrame:
    """
    Read a CSV UploadedFile and normalise all column names to lowercase_underscore.

    Returns:
        pd.DataFrame with normalised columns.

    Raises:
        ValueError: if the file cannot be parsed as CSV.
    """
    df = pd.read_csv(uploaded_file)
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    return df


# ──────────────────────────────────────────────────────────────────────────────
# SCORE
# ──────────────────────────────────────────────────────────────────────────────

def score_candidates(df: pd.DataFrame, jd: dict) -> tuple[pd.DataFrame, str | None]:
    """
    Score and rank every candidate against the parsed JD.

    Args:
        df:  Candidate DataFrame (normalised columns).
        jd:  Dict produced by jd_parser.extract_jd_with_ai().

    Returns:
        (scored_df, name_col)
        scored_df  — original DataFrame + score columns, sorted descending by total_score
        name_col   — detected column name for candidate names (or None)
    """
    df = df.copy()

    # JD fields
    required_skills = [s.lower().strip() for s in jd.get("skills", [])]
    jd_role         = jd.get("role", "").lower()
    jd_location     = jd.get("location", "").lower()
    exp_min         = int(jd.get("experience_min", 0) or 0)
    exp_max         = int(jd.get("experience_max", 99) or 99)

    # Detect columns
    skills_col  = _find_col(df, "skills")
    role_col    = _find_col(df, "role")
    location_col= _find_col(df, "location")
    exp_col     = _find_col(df, "experience")
    name_col    = _find_col(df, "name")

    rows_skill_sc, rows_role_sc, rows_exp_sc = [], [], []
    rows_loc, rows_matched, rows_missing, rows_total = [], [], [], []

    for _, row in df.iterrows():

        # ── Skill Score (0–40) ────────────────────────────────────────────────
        cand_skills_raw = str(row[skills_col]).lower() if skills_col else ""
        matched = [s for s in required_skills if s in cand_skills_raw]
        missing = [s for s in required_skills if s not in cand_skills_raw]
        skill_sc = round(len(matched) / max(len(required_skills), 1) * 40)

        # ── Role Score (0–30) ─────────────────────────────────────────────────
        cand_role = str(row[role_col]).lower() if role_col else ""
        if jd_role and cand_role:
            jd_words   = set(jd_role.split())
            role_words = set(cand_role.split())
            overlap    = jd_words & role_words
            role_sc    = round(len(overlap) / max(len(jd_words), 1) * 30)
        else:
            role_sc = 0

        # ── Experience Score (0–30) ───────────────────────────────────────────
        cand_exp = _parse_experience(row[exp_col]) if exp_col else 0.0
        if exp_min <= cand_exp <= exp_max:
            exp_sc = 30
        elif cand_exp > exp_max:
            exp_sc = max(0, 30 - int((cand_exp - exp_max) * 3))
        else:
            exp_sc = max(0, 30 - int((exp_min - cand_exp) * 5))

        # ── Location Match (boolean) ──────────────────────────────────────────
        cand_loc  = str(row[location_col]).lower() if location_col else ""
        loc_match = bool(jd_location and jd_location in cand_loc)

        total = skill_sc + role_sc + exp_sc

        rows_skill_sc.append(skill_sc)
        rows_role_sc.append(role_sc)
        rows_exp_sc.append(exp_sc)
        rows_loc.append("✅ Yes" if loc_match else "❌ No")
        rows_matched.append(", ".join(matched) if matched else "None")
        rows_missing.append(", ".join(missing) if missing else "None")
        rows_total.append(total)

    df["total_score"]      = rows_total
    df["skill_score"]      = rows_skill_sc
    df["role_score"]       = rows_role_sc
    df["experience_score"] = rows_exp_sc
    df["location_match"]   = rows_loc
    df["matched_skills"]   = rows_matched
    df["missing_skills"]   = rows_missing

    df = df.sort_values("total_score", ascending=False).reset_index(drop=True)
    df.insert(0, "rank", range(1, len(df) + 1))

    return df, name_col


# ──────────────────────────────────────────────────────────────────────────────
# EXPORT — EXCEL
# ──────────────────────────────────────────────────────────────────────────────

def export_excel(df: pd.DataFrame, jd: dict) -> bytes:
    """
    Build a colour-coded Excel workbook with two sheets:
      Sheet 1 — Ranked Candidates (green ≥70 · yellow ≥45 · red <45)
      Sheet 2 — JD Summary

    Returns:
        bytes: Raw .xlsx file content ready for st.download_button.
    """
    wb = Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_fill  = PatternFill("solid", fgColor="6366F1")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin",   color="D1D5DB"),
        right=Side(style="thin",  color="D1D5DB"),
        top=Side(style="thin",    color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    fill_green  = PatternFill("solid", fgColor="D1FAE5")  # score ≥ 70
    fill_yellow = PatternFill("solid", fgColor="FEF3C7")  # score ≥ 45
    fill_red    = PatternFill("solid", fgColor="FEE2E2")  # score < 45

    # ── Sheet 1: Ranked Candidates ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Ranked Candidates"

    # Title banner
    ws.merge_cells("A1:M1")
    ws["A1"] = f"🎯 AI Recruiter — Results for: {jd.get('role', 'Role')}"
    ws["A1"].font  = Font(bold=True, size=14, color="1E293B", name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill  = PatternFill("solid", fgColor="F1F5F9")
    ws.row_dimensions[1].height = 30

    headers = [
        "Rank", "Name", "Role", "Location", "Experience (yrs)",
        "Skills", "Total Score", "Skill Score", "Role Score",
        "Exp Score", "Location Match", "Matched Skills", "Missing Skills",
    ]
    col_widths = [6, 22, 22, 14, 16, 32, 13, 13, 12, 12, 16, 30, 30]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[2].height = 24

    # Map logical keys → actual df column names (graceful fallback)
    key_map = [
        "rank", "name", "role", "location", "experience",
        "skills", "total_score", "skill_score", "role_score",
        "experience_score", "location_match", "matched_skills", "missing_skills",
    ]
    df_cols = set(df.columns)

    def _get(row, key):
        return row[key] if key in df_cols else ""

    for r_idx, (_, row) in enumerate(df.iterrows(), start=3):
        total = _get(row, "total_score")
        row_fill = fill_green if total >= 70 else (fill_yellow if total >= 45 else fill_red)

        for c_idx, key in enumerate(key_map, start=1):
            val  = _get(row, key)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if key == "total_score":
                cell.fill = row_fill
                cell.font = Font(bold=True, name="Calibri", size=11)
            elif key == "rank" and val in (1, 2, 3):
                cell.font = Font(bold=True, color="4F46E5", name="Calibri")

        ws.row_dimensions[r_idx].height = 18

    ws.freeze_panes = "A3"

    # ── Sheet 2: JD Summary ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("JD Summary")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 55

    ws2.merge_cells("A1:B1")
    ws2["A1"] = "📋 Job Description — Extracted Details"
    ws2["A1"].font      = Font(bold=True, size=13, color="1E293B", name="Calibri")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["A1"].fill      = PatternFill("solid", fgColor="EDE9FE")
    ws2.row_dimensions[1].height = 28

    jd_fields = [
        ("Role",            jd.get("role", "")),
        ("Location",        jd.get("location", "")),
        ("Experience Min",  f"{jd.get('experience_min', '')} years"),
        ("Experience Max",  f"{jd.get('experience_max', '')} years"),
        ("Industry",        jd.get("industry", "")),
        ("Employment Type", jd.get("employment_type", "")),
        ("Education",       jd.get("education", "") or "Not specified"),
        ("Company",         jd.get("company", "")  or "Not specified"),
        ("Required Skills", ", ".join(jd.get("skills", []))),
        ("Summary",         jd.get("summary", "")),
    ]

    lbl_font = Font(bold=True, name="Calibri", color="4F46E5")
    for i, (label, value) in enumerate(jd_fields, start=2):
        ws2.cell(row=i, column=1, value=label).font = lbl_font
        ws2.cell(row=i, column=2, value=value)
        ws2.row_dimensions[i].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# EXPORT — CSV
# ──────────────────────────────────────────────────────────────────────────────

def export_csv(df: pd.DataFrame) -> bytes:
    """Return the scored DataFrame as UTF-8 CSV bytes."""
    return df.to_csv(index=False).encode("utf-8")


# ──────────────────────────────────────────────────────────────────────────────
# QUICK SELF-TEST  (run: python filter.py)
# ──────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import io as _io

    sample_csv = """name,role,location,experience,skills
Priya Sharma,Data Scientist,Mumbai,4,Python ML SQL Pandas Sklearn
Rohit Verma,ML Engineer,Mumbai,3,Python ML TensorFlow NLP
Arjun Rao,Business Analyst,Delhi,5,Excel SQL Tableau
"""
    sample_jd = {
        "role": "Data Scientist",
        "skills": ["python", "ml", "sql", "pandas"],
        "location": "mumbai",
        "experience_min": 3,
        "experience_max": 6,
    }

    df_in = pd.read_csv(_io.StringIO(sample_csv))
    df_in.columns = [c.strip().lower().replace(" ", "_") for c in df_in.columns]
    scored, name_col = score_candidates(df_in, sample_jd)
    print(scored[["rank", "name", "total_score", "skill_score", "role_score", "experience_score"]].to_string(index=False))
    print(f"\nName column detected: {name_col}")
